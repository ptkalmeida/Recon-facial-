import io
from datetime import datetime
from typing import List, Dict, Any

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False


def generate_excel_report(data: List[Dict[str, Any]], report_type: str) -> bytes:
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl não está instalado")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório"
    
    headers = list(data[0].keys()) if data else []
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    for row_num, row_data in enumerate(data, 2):
        for col_num, header in enumerate(headers, 1):
            value = row_data.get(header, "")
            ws.cell(row=row_num, column=col_num, value=value)
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_pdf_report(data: List[Dict[str, Any]], report_type: str) -> bytes:
    if not HAS_REPORTLAB:
        raise ImportError("reportlab não está instalado")
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    
    title_text = f"Relatório de {report_type.replace('_', ' ').title()}"
    title = Paragraph(title_text, styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 20))
    
    date_text = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    date_paragraph = Paragraph(date_text, styles['Normal'])
    elements.append(date_paragraph)
    elements.append(Spacer(1, 20))
    
    if data:
        headers = list(data[0].keys())
        table_data = [headers]
        
        for row in data[:50]:
            table_data.append([str(row.get(h, "")) for h in headers])
        
        table = Table(table_data)
        
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ])
        
        table.setStyle(style)
        elements.append(table)
    
    doc.build(elements)
    return buffer.getvalue()